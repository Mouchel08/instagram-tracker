"""
Export all Instagram analytics to a formatted Excel tracker.
Run: python export_excel.py
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from analyze import (
    by_content_type,
    caption_length_analysis,
    hashtag_performance,
    load_data,
    overview_metrics,
    posting_frequency,
    strategy_recommendations,
    best_posting_times,
    top_posts,
)
from config import DATA_DIR


# ── Colour palette ─────────────────────────────────────────────────────────

DARK_BG    = "1A1A2E"
MID_BG     = "16213E"
ACCENT     = "E94560"
ACCENT2    = "0F3460"
TEAL       = "2EC4B6"
TEXT_WHITE = "FFFFFF"
TEXT_GREY  = "AAAAAA"
LIGHT_GREY = "F5F5F5"
ALT_ROW    = "F0F4FF"

HEADER_FILL   = PatternFill("solid", fgColor=DARK_BG)
SUBHEAD_FILL  = PatternFill("solid", fgColor=ACCENT2)
ACCENT_FILL   = PatternFill("solid", fgColor=ACCENT)
ALT_FILL      = PatternFill("solid", fgColor=ALT_ROW)
TEAL_FILL     = PatternFill("solid", fgColor=TEAL)

HEADER_FONT   = Font(name="Calibri", bold=True, color=TEXT_WHITE, size=11)
SUBHEAD_FONT  = Font(name="Calibri", bold=True, color=TEXT_WHITE, size=10)
BODY_FONT     = Font(name="Calibri", size=10)
TITLE_FONT    = Font(name="Calibri", bold=True, color=TEXT_WHITE, size=14)

THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def style_header_row(ws, row: int, fill=None):
    fill = fill or HEADER_FILL
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER


def style_data_row(ws, row: int, alt: bool = False):
    for cell in ws[row]:
        cell.font = BODY_FONT
        cell.alignment = LEFT if cell.column == 1 else CENTER
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def freeze_and_filter(ws, freeze="A2", filter_range=None):
    ws.freeze_panes = freeze
    if filter_range:
        ws.auto_filter.ref = filter_range


# ── Sheet 1: Summary ───────────────────────────────────────────────────────

def write_summary(ws, overview: dict, account: dict):
    ws.title = "Summary"
    ws.sheet_properties.tabColor = ACCENT

    # Title block
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Instagram Analytics — @{account.get('username', '')}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BG)
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    ws["A2"].value = (
        f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}  ·  "
        f"Data: {overview['date_range_start']} to {overview['date_range_end']}"
    )
    ws["A2"].font = Font(name="Calibri", italic=True, color=TEXT_GREY, size=10)
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 20

    # KPI header
    ws.merge_cells("A4:F4")
    ws["A4"].value = "KEY METRICS"
    ws["A4"].font = SUBHEAD_FONT
    ws["A4"].fill = SUBHEAD_FILL
    ws["A4"].alignment = CENTER
    ws.row_dimensions[4].height = 24

    kpis = [
        ("Metric", "Value", "", "Metric", "Value", ""),
        ("Total Posts", overview["total_posts"], "", "Avg Engagement Rate", f"{overview['avg_engagement_rate']:.2f}%", ""),
        ("Total Likes", f"{overview['total_likes']:,}", "", "Median Engagement Rate", f"{overview['median_engagement_rate']:.2f}%", ""),
        ("Total Comments", f"{overview['total_comments']:,}", "", "Avg Likes / Post", f"{overview['avg_likes']:,.1f}", ""),
        ("Total Saves", f"{overview['total_saves']:,}", "", "Avg Comments / Post", f"{overview['avg_comments']:,.1f}", ""),
        ("Total Shares", f"{overview['total_shares']:,}", "", "Avg Saves / Post", f"{overview['avg_saves']:,.1f}", ""),
        ("Total Impressions", f"{overview['total_impressions']:,}", "", "Followers", f"{overview['followers']:,}", ""),
        ("Total Reach", f"{overview['total_reach']:,}", "", "Following", f"{overview['following']:,}", ""),
    ]

    for i, row_data in enumerate(kpis, start=5):
        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if i == 5:
                cell.fill = SUBHEAD_FILL
                cell.font = SUBHEAD_FONT
                cell.alignment = CENTER
            else:
                cell.font = Font(name="Calibri", bold=(j in (1, 4)), size=10)
                cell.fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                cell.alignment = CENTER if j in (2, 5) else LEFT
            cell.border = THIN_BORDER
        ws.row_dimensions[i].height = 22

    set_col_widths(ws, {"A": 28, "B": 20, "C": 4, "D": 28, "E": 20, "F": 4})


# ── Sheet 2: All Posts ─────────────────────────────────────────────────────

def write_all_posts(ws, df: pd.DataFrame):
    ws.title = "All Posts"
    ws.sheet_properties.tabColor = ACCENT2

    columns = [
        ("Date", "date", 14),
        ("Type", "media_type", 12),
        ("Likes", "likes", 10),
        ("Comments", "comments", 10),
        ("Saves", "saves", 10),
        ("Shares", "shares", 10),
        ("Reach", "reach", 12),
        ("Impressions", "impressions", 13),
        ("Engagement %", "engagement_rate", 14),
        ("Save Rate %", "save_rate", 12),
        ("Reach Rate %", "reach_rate", 13),
        ("Hashtag #", "hashtag_count", 11),
        ("Caption Len", "caption_length", 12),
        ("Day", "day_of_week", 12),
        ("Hour", "hour", 8),
        ("Hashtags", "hashtags", 40),
        ("Caption", "caption", 60),
        ("Link", "permalink", 50),
    ]

    headers = [c[0] for c in columns]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = columns[j-1][2]

    ws.row_dimensions[1].height = 24

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        alt = (i % 2 == 0)
        for j, (_, col_key, _) in enumerate(columns, start=1):
            val = row.get(col_key, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
            if col_key in ("engagement_rate", "save_rate", "reach_rate"):
                cell.number_format = "0.00\"%\""
            elif col_key in ("likes", "comments", "saves", "shares", "reach", "impressions"):
                cell.number_format = "#,##0"
            cell.alignment = LEFT if col_key in ("caption", "hashtags", "permalink") else CENTER
        ws.row_dimensions[i].height = 18

    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(columns))}1")


# ── Sheet 3: Content Type breakdown ───────────────────────────────────────

def write_by_type(ws, types: pd.DataFrame):
    ws.title = "By Content Type"
    ws.sheet_properties.tabColor = "533483"

    headers = [
        "Type", "Posts", "Avg ER %", "Avg Likes", "Avg Comments",
        "Avg Saves", "Avg Shares", "Avg Reach", "Avg Impressions",
        "Total Likes", "Total Comments",
    ]
    keys = [
        "media_type", "post_count", "avg_engagement_rate", "avg_likes",
        "avg_comments", "avg_saves", "avg_shares", "avg_reach",
        "avg_impressions", "total_likes", "total_comments",
    ]
    widths = [16, 10, 14, 12, 15, 12, 12, 12, 16, 14, 16]

    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, (_, row) in enumerate(types.iterrows(), start=2):
        for j, key in enumerate(keys, start=1):
            val = row.get(key, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            if key == "avg_engagement_rate":
                cell.number_format = "0.00\"%\""
        ws.row_dimensions[i].height = 22

    ws.row_dimensions[1].height = 24
    freeze_and_filter(ws, "A2")


# ── Sheet 4: Top Posts ─────────────────────────────────────────────────────

def write_top_posts(ws, top: pd.DataFrame):
    ws.title = "Top 20 Posts"
    ws.sheet_properties.tabColor = ACCENT

    cols = ["date", "media_type", "likes", "comments", "saves", "shares",
            "reach", "engagement_rate", "hashtags", "caption", "permalink"]
    headers = ["Date", "Type", "Likes", "Comments", "Saves", "Shares",
               "Reach", "ER %", "Hashtags", "Caption", "Link"]
    widths = [14, 10, 10, 10, 10, 10, 12, 10, 40, 60, 50]

    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = ACCENT_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = w

    top_filtered = top[[c for c in cols if c in top.columns]]
    for i, (_, row) in enumerate(top_filtered.iterrows(), start=2):
        for j, col in enumerate([c for c in cols if c in top.columns], start=1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT if col in ("caption", "hashtags", "permalink") else CENTER
            cell.border = THIN_BORDER
            cell.fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            if col == "engagement_rate":
                cell.number_format = "0.00\"%\""
        ws.row_dimensions[i].height = 18

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


# ── Sheet 5: Hashtags ──────────────────────────────────────────────────────

def write_hashtags(ws, tags: pd.DataFrame):
    ws.title = "Hashtag Performance"
    ws.sheet_properties.tabColor = "2EC4B6"

    headers = ["Hashtag", "Uses", "Avg ER %", "Avg Likes", "Avg Comments", "Avg Saves", "Avg Reach"]
    keys    = ["hashtag", "uses", "avg_engagement_rate", "avg_likes", "avg_comments", "avg_saves", "avg_reach"]
    widths  = [24, 8, 14, 12, 15, 12, 12]

    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = PatternFill("solid", fgColor="2EC4B6")
        cell.font = Font(name="Calibri", bold=True, color="1A1A2E", size=11)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, (_, row) in enumerate(tags.iterrows(), start=2):
        for j, key in enumerate(keys, start=1):
            val = row.get(key, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT if key == "hashtag" else CENTER
            cell.border = THIN_BORDER
            cell.fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            if key == "avg_engagement_rate":
                cell.number_format = "0.00\"%\""
        ws.row_dimensions[i].height = 18

    ws.row_dimensions[1].height = 24
    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")


# ── Sheet 6: Timing ────────────────────────────────────────────────────────

def write_timing(ws, timing: dict):
    ws.title = "Timing"
    ws.sheet_properties.tabColor = "FF9F1C"

    # Day section
    ws.merge_cells("A1:C1")
    ws["A1"].value = "Best Days to Post"
    ws["A1"].font = SUBHEAD_FONT
    ws["A1"].fill = SUBHEAD_FILL
    ws["A1"].alignment = CENTER

    for j, h in enumerate(["Day", "Avg ER %", "Post Count"], start=1):
        ws.cell(row=2, column=j, value=h).fill = HEADER_FILL
        ws.cell(row=2, column=j).font = HEADER_FONT
        ws.cell(row=2, column=j).alignment = CENTER
        ws.cell(row=2, column=j).border = THIN_BORDER

    for i, (_, row) in enumerate(timing["by_day"].iterrows(), start=3):
        for j, key in enumerate(["day_of_week", "avg_engagement_rate", "post_count"], start=1):
            cell = ws.cell(row=i, column=j, value=row.get(key, ""))
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if key == "avg_engagement_rate":
                cell.number_format = "0.00\"%\""

    # Hour section
    start_col = 5
    ws.cell(row=1, column=start_col, value="Best Hours to Post").fill = PatternFill("solid", fgColor="FF9F1C")
    ws.cell(row=1, column=start_col).font = Font(name="Calibri", bold=True, color="1A1A2E", size=11)
    ws.cell(row=1, column=start_col).alignment = CENTER

    for j, h in enumerate(["Hour", "Avg ER %", "Post Count"], start=start_col):
        ws.cell(row=2, column=j, value=h).fill = HEADER_FILL
        ws.cell(row=2, column=j).font = HEADER_FONT
        ws.cell(row=2, column=j).alignment = CENTER

    for i, (_, row) in enumerate(timing["by_hour"].iterrows(), start=3):
        for j, key in enumerate(["hour", "avg_engagement_rate", "post_count"], start=start_col):
            val = row.get(key, "")
            if key == "hour" and val != "":
                val = f"{int(val):02d}:00"
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            if key == "avg_engagement_rate":
                cell.number_format = "0.00\"%\""

    set_col_widths(ws, {"A": 16, "B": 14, "C": 12, "D": 4, "E": 10, "F": 14, "G": 12})


# ── Sheet 7: Strategy ──────────────────────────────────────────────────────

def write_strategy(ws, recs: list[dict]):
    ws.title = "Strategy"
    ws.sheet_properties.tabColor = ACCENT

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Data-Driven Strategy Recommendations"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BG)
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    for j, (h, w) in enumerate(zip(["Category", "Insight", "Action", "Notes"], [18, 55, 55, 30]), start=1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.fill = ACCENT_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, rec in enumerate(recs, start=3):
        ws.cell(row=i, column=1, value=rec["category"]).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=i, column=2, value=rec["insight"]).font = BODY_FONT
        ws.cell(row=i, column=3, value=rec["action"]).font = BODY_FONT
        ws.cell(row=i, column=4, value="").font = BODY_FONT
        for j in range(1, 5):
            ws.cell(row=i, column=j).border = THIN_BORDER
            ws.cell(row=i, column=j).alignment = LEFT
            ws.cell(row=i, column=j).fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        ws.row_dimensions[i].height = 36

    ws.row_dimensions[2].height = 24


# ── Sheet 8: Monthly ───────────────────────────────────────────────────────

def write_monthly(ws, freq: pd.DataFrame):
    ws.title = "Monthly Breakdown"
    ws.sheet_properties.tabColor = "533483"

    headers = ["Month", "Posts", "Avg ER %", "Avg Likes", "Avg Saves"]
    keys    = ["year_month", "post_count", "avg_engagement_rate", "avg_likes", "avg_saves"]
    widths  = [14, 10, 14, 12, 12]

    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, (_, row) in enumerate(freq.iterrows(), start=2):
        for j, key in enumerate(keys, start=1):
            val = row.get(key, "")
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            if key == "avg_engagement_rate":
                cell.number_format = "0.00\"%\""
        ws.row_dimensions[i].height = 18

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


# ── Main ───────────────────────────────────────────────────────────────────

def export_to_excel(output_path: str = None) -> str:
    df, account = load_data()
    from analyze import run_full_analysis
    analysis = run_full_analysis(df, account)

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"instagram_tracker_{account.get('username', 'export')}_{ts}.xlsx"
        output_path = os.path.join(DATA_DIR, filename)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write placeholder frames to create sheets
        pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="All Posts", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="By Content Type", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Top 20 Posts", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Hashtag Performance", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Timing", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Strategy", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Monthly Breakdown", index=False)

    wb = load_workbook(output_path)

    write_summary(wb["Summary"], analysis["overview"], account)
    write_all_posts(wb["All Posts"], df)
    write_by_type(wb["By Content Type"], analysis["by_type"])
    write_top_posts(wb["Top 20 Posts"], analysis["top_posts"])
    write_hashtags(wb["Hashtag Performance"], analysis["hashtags"])
    write_timing(wb["Timing"], analysis["timing"])
    write_strategy(wb["Strategy"], analysis["recommendations"])
    write_monthly(wb["Monthly Breakdown"], analysis["frequency"])

    wb.save(output_path)
    print(f"Excel tracker saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    path = export_to_excel()
    print(f"Done: {path}")
